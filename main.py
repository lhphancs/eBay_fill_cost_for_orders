import openpyxl
import csv
import os
from pathlib import Path

import getUserInput
import inputMsgs

def getReadExcelPath(readDirPath):
    readExcelPath = None
    for file in os.listdir(readDirPath):
        if file.endswith('xlsx'):
            readExcelPath = os.path.join(readDirPath, file)
            break
    return readExcelPath

def getNameToSheetDict(readWb):
    retDict = {}
    for sheet in readWb:
        retDict[sheet.title] = sheet
    return retDict

"""This also allows 'must not contain'
by entering phrase of '! Apple cider', this means title must not contain Apple cider"""
def meetsAllPhraseConditions(title, phrases):
    uppercasedTitle = title.upper()
    for phrase in phrases:
        uppercasedPhrase = phrase.upper()
        splittedPhrase = uppercasedPhrase.split(' ', 1)
        if splittedPhrase[0] == '!' and len(splittedPhrase) > 1:
            if splittedPhrase[1] in uppercasedTitle:
                return False
        elif uppercasedPhrase not in uppercasedTitle:
            return False
    return True

def getDictOfMatches(orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict, phrases):
    retDict = {}
    for key in orderedSheetNames:
        matchingRows = []
        sheet = nameToSheetDict[key]
        nameToTitleAndCostLocation = nameToTitleAndCostLocationDict[sheet.title]
        titleCol = nameToTitleAndCostLocation[0][1]
        for rowIndex in range(nameToTitleAndCostLocation[0][0] + 1, sheet.max_row + 1):
            title = sheet.cell(rowIndex, titleCol).value
            if title != None and meetsAllPhraseConditions(title, phrases):
                matchingRows.append(rowIndex)
        retDict[key] = matchingRows
    return retDict

def printResultOfFind(orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict, dictOfMatches):
    for key in orderedSheetNames:
        titleCol = nameToTitleAndCostLocationDict[key][0][1]
        print( str.format('{}: ', key) )
        for row in dictOfMatches[key]:
            print( '\t' + nameToSheetDict[key].cell(row, titleCol).value)

def getTitlePhrasesAndPrintFinds(orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict):
    phrases = getUserInput.getPhrases()
    dictOfMatches = getDictOfMatches(orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict, phrases)
    printResultOfFind(orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict, dictOfMatches)
    return dictOfMatches


def writeCostToCell(orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict, dictOfMatches, cost):
    writeAmt = 0
    for key in orderedSheetNames:
        writeAmt = writeAmt + len( dictOfMatches[key] )
        sheet = nameToSheetDict[key]
        costCol = nameToTitleAndCostLocationDict[key][1][1]
        for row in dictOfMatches[key]:
            sheet.cell(row, costCol).value = cost
    return writeAmt

def handlePhraseWriteCommand(orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict):
    dictOfMatches = getTitlePhrasesAndPrintFinds(orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict)
    cost = getUserInput.getCost()
    confirmResp = getUserInput.getConfirmation("Are you sure you want to write? (Y)es or (N)o: ")
    if confirmResp == 'Y':
        return writeCostToCell(orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict, dictOfMatches, cost)

    else:
        return 0

def getCost(sheet, row, costCol):
    cost = sheet.cell(row, costCol).value
    if cost != None:
        try:
            return float(cost)
        except ValueError:
            print( str.format('!!!!!!!!!!!!!!!!!!! Invalid cost detected at row {} !!!!!!!!!!!!!!!!!!!', row) )
    return None

def getListingFinalSellPrice(phraseAfterSemiColon):
    splitOnFinalPrice = phraseAfterSemiColon.split('FINAL PRICE: $')
    if len(splitOnFinalPrice) == 1:
        return None
    finalSellPrice = splitOnFinalPrice[1].split()[0].strip(',')
    return float(finalSellPrice)

def getTitleAndFinalSellPriceTupleOrNone(sheet, row, titleCol):
    titleCellVal = sheet.cell(row, titleCol).value
    if titleCellVal == None:
        return None
    titleCellVal = titleCellVal.upper()
    rsplitValsWithSemicolon = titleCellVal.rsplit(';', 1)
    if len(rsplitValsWithSemicolon) != 2:
        return None
    listingTitle = rsplitValsWithSemicolon[0].strip("'")
    finalSellPrice = getListingFinalSellPrice(rsplitValsWithSemicolon[1])
    if listingTitle == None or finalSellPrice == None:
        return None
    return (listingTitle, finalSellPrice)

def writeCommons(orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict, combinedStr, cost):
    for key in orderedSheetNames:
        titleCol = nameToTitleAndCostLocationDict[key][0][1]
        costCol = nameToTitleAndCostLocationDict[key][1][1]
        sheet = nameToSheetDict[key]
        for rowIndex in range(nameToTitleAndCostLocationDict[key][0][0] + 1, sheet.max_row + 1):
            titleAndFinalPriceTuple = getTitleAndFinalSellPriceTupleOrNone(sheet, rowIndex, titleCol)
            if titleAndFinalPriceTuple != None:
                if combinedStr == getCombinedStr(titleAndFinalPriceTuple):
                    sheet.cell(rowIndex, costCol).value = cost

def getCombinedStr(titleAndFinalPriceTuple):
    return titleAndFinalPriceTuple[0] + ' $' + str(titleAndFinalPriceTuple[1])

def handleCommonsWriting(orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict):
    encountersSet = set()
    for key in orderedSheetNames:
        print( str.format('=================================== {} ===================================', key) )
        titleCol = nameToTitleAndCostLocationDict[key][0][1]
        costCol = nameToTitleAndCostLocationDict[key][1][1]
        sheet = nameToSheetDict[key]
        for rowIndex in range(nameToTitleAndCostLocationDict[key][0][0] + 1, sheet.max_row + 1):
            titleAndFinalPriceTuple = getTitleAndFinalSellPriceTupleOrNone(sheet, rowIndex, titleCol)
            cost = getCost(sheet, rowIndex, costCol)
            if titleAndFinalPriceTuple != None and cost != None:
                combinedStr = getCombinedStr(titleAndFinalPriceTuple)
                if (combinedStr not in encountersSet):
                    tTitle = titleAndFinalPriceTuple[0].ljust(80)
                    tEbayFinalPrice = str(titleAndFinalPriceTuple[1]).ljust(5)
                    print( str.format('{}     ${}     myCost = ${}', tTitle, tEbayFinalPrice, cost) )
                    encountersSet.add(combinedStr)
                    writeCommons(orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict, combinedStr, cost)
        print()

def saveFile(excelWb, outputPath):
    while True:
        try:
            excelWb.save(outputPath)
            print("Save successful.")
            break
        except PermissionError as e:
            print(e)
            print("Is the write file, 'result.xlsx', open?")
            print("Press enter to try again...")
            input('(Note: Changes made to the excel file outside of this program will not be saved)\n')
        
def handleCommand(excelWb, outputPath, orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict, uppercasedCommand):
    if uppercasedCommand == 'F':
        getTitlePhrasesAndPrintFinds(orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict)
    elif uppercasedCommand == 'P':
        handlePhraseWriteCommand(orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict)
        
    elif uppercasedCommand == 'C':
        if 'Y' == getUserInput.getConfirmation("Are you sure you want to write ALL commons? (Y)es or (N)o: "):
            handleCommonsWriting(orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict)
        
    elif uppercasedCommand == 'H':
        print(inputMsgs.menuHelp) 

    elif uppercasedCommand == 'E':
        pass
    else:
        print('Invalid command. Try again.')
    if uppercasedCommand == 'P' or uppercasedCommand == 'C':
        saveFile(excelWb, outputPath)
    print()

def getLocationOfWord(sheet, word):
    MAX_INDEX_TO_ITER = 20
    for rowIndex in range(1, MAX_INDEX_TO_ITER + 1):
        for colIndex in range(1, MAX_INDEX_TO_ITER + 1):
            cellVal = sheet.cell(rowIndex, colIndex).value
            if cellVal != None and type(cellVal) is str:
                cellVal = cellVal.strip(" '")
                if cellVal.upper() == word.upper():
                    return (rowIndex, colIndex)
    return None

def getNameToTitleAndCostLocationDict(excelWb):
    nameToTitleLocationDict = {}
    for sheet in excelWb:
        titleLocation = getLocationOfWord(sheet, 'TITLE')
        costLocation = getLocationOfWord(sheet, 'COST')
        if titleLocation == None:
            print( str.format("Error...Could not find 'Title' header in sheet: {}", sheet.title) )
            return None
        elif costLocation == None:
            print( str.format("Error...Could not find 'Cost' header in sheet: {}", sheet.title) )
            return None
        else:
            nameToTitleLocationDict[sheet.title] = (titleLocation, costLocation)
        
    return nameToTitleLocationDict

def handlePromptAndResponses(excelWb, outputPath):
    orderedSheetNames = excelWb.get_sheet_names()
    nameToSheetDict = getNameToSheetDict(excelWb)
    nameToTitleAndCostLocationDict = getNameToTitleAndCostLocationDict(excelWb)
    if nameToTitleAndCostLocationDict != None:
        uppercasedCommand = None
        while uppercasedCommand != 'E':
            uppercasedCommand = getUserInput.getMenuCmd()
            handleCommand(excelWb, outputPath, orderedSheetNames, nameToSheetDict, nameToTitleAndCostLocationDict, uppercasedCommand)

if __name__ == '__main__':
    curDir = os.path.dirname(os.path.abspath(__file__))
    readDirPath = os.path.join(curDir, 'placeSingleExcelFileHere')
    readExcelPath = getReadExcelPath(readDirPath)
    outputPath = os.path.join(curDir, 'result', 'result.xlsx')

    if readExcelPath != None:
        excelWb = openpyxl.load_workbook(readExcelPath, data_only=True)
        handlePromptAndResponses(excelWb, outputPath)
    else:
        print( str.format("No file with extension 'xlsx' in: {}", readDirPath) )
    print("Program has finished...")